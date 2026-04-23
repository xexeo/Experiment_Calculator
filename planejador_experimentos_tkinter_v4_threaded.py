
import json
import itertools
import random
import time
import threading
import queue
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


APP_TITLE = "Planejador de Experimentos Discretos"
FULL_ENUMERATION_LIMIT = 50000
DEFAULT_CANDIDATE_POOL = 6000
DEFAULT_ITERATIONS = 250
RANDOM_SEED = 42


@dataclass
class Hyperparameter:
    name: str
    values: list


class CancelledError(Exception):
    pass


def normalize_values_text(text: str):
    parts = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if ";" in line:
            items = [x.strip() for x in line.split(";")]
        elif "," in line:
            items = [x.strip() for x in line.split(",")]
        else:
            items = [line]
        for item in items:
            if item:
                parts.append(item)
    seen = set()
    out = []
    for item in parts:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def full_factorial_count(hparams):
    total = 1
    for hp in hparams:
        total *= len(hp.values)
    return total


def generate_full_candidates(hparams):
    names = [hp.name for hp in hparams]
    value_lists = [hp.values for hp in hparams]
    return [dict(zip(names, combo)) for combo in itertools.product(*value_lists)]


def deduplicate_configs(configs, ordered_names):
    seen = set()
    out = []
    for cfg in configs:
        key = tuple(cfg[name] for name in ordered_names)
        if key not in seen:
            seen.add(key)
            out.append(cfg)
    return out


def random_config(hparams, rng):
    return {hp.name: rng.choice(hp.values) for hp in hparams}


def generate_seed_candidates(hparams):
    base = {hp.name: hp.values[0] for hp in hparams}
    seeds = []
    for hp in hparams:
        for value in hp.values:
            cfg = dict(base)
            cfg[hp.name] = value
            seeds.append(cfg)
    return deduplicate_configs(seeds, [hp.name for hp in hparams])


def generate_candidate_pool(hparams, pool_size, rng, cancel_check=None):
    total = full_factorial_count(hparams)
    if total <= FULL_ENUMERATION_LIMIT:
        return generate_full_candidates(hparams), True
    names = [hp.name for hp in hparams]
    candidates = generate_seed_candidates(hparams)
    seen = {tuple(cfg[name] for name in names) for cfg in candidates}
    target = max(pool_size, len(candidates))
    attempts = 0
    max_attempts = target * 40 + 1000
    while len(candidates) < target and attempts < max_attempts:
        if cancel_check:
            cancel_check()
        cfg = random_config(hparams, rng)
        key = tuple(cfg[name] for name in names)
        if key not in seen:
            seen.add(key)
            candidates.append(cfg)
        attempts += 1
    return candidates, False


def hamming_distance(cfg_a, cfg_b, names):
    return sum(1 for name in names if cfg_a[name] != cfg_b[name])


def build_twise_universe(hparams, t):
    universe = {}
    if t < 2 or t > len(hparams):
        return universe
    for combo in itertools.combinations(range(len(hparams)), t):
        hp_names = tuple(hparams[i].name for i in combo)
        values_product = itertools.product(*[hparams[i].values for i in combo])
        universe[hp_names] = set(values_product)
    return universe


def covered_twise_of_config(cfg, hparams, t):
    covered = []
    if t < 2 or t > len(hparams):
        return covered
    for combo in itertools.combinations(range(len(hparams)), t):
        hp_names = tuple(hparams[i].name for i in combo)
        hp_values = tuple(cfg[hparams[i].name] for i in combo)
        covered.append((hp_names, hp_values))
    return covered


def compute_targets(hparams, unique_experiments):
    marginal_targets = {}
    for hp in hparams:
        target = unique_experiments / max(1, len(hp.values))
        for val in hp.values:
            marginal_targets[(hp.name, val)] = target
    return marginal_targets


def compute_selected_counts(selected, hparams, t_strength):
    marginal_counts = Counter()
    twise_counts = Counter()
    twise_presence = defaultdict(set)
    names = [hp.name for hp in hparams]
    for cfg in selected:
        for hp in hparams:
            marginal_counts[(hp.name, cfg[hp.name])] += 1
        for tw_key, tw_val in covered_twise_of_config(cfg, hparams, t_strength):
            twise_counts[(tw_key, tw_val)] += 1
            twise_presence[tw_key].add(tw_val)
    min_dist = 0
    avg_dist = 0.0
    if len(selected) >= 2:
        dists = []
        for i in range(len(selected)):
            for j in range(i + 1, len(selected)):
                dists.append(hamming_distance(selected[i], selected[j], names))
        if dists:
            min_dist = min(dists)
            avg_dist = sum(dists) / len(dists)
    return marginal_counts, twise_counts, twise_presence, min_dist, avg_dist


def score_candidate(cfg, selected, hparams, t_strength, twise_presence, marginal_counts, marginal_targets, weights):
    names = [hp.name for hp in hparams]
    score = 0.0

    if weights["twise"] > 0:
        new_twise = 0
        for tw_key, tw_val in covered_twise_of_config(cfg, hparams, t_strength):
            if tw_val not in twise_presence[tw_key]:
                new_twise += 1
        score += weights["twise"] * new_twise

    if weights["marginal"] > 0:
        marginal_gain = 0.0
        for hp in hparams:
            key = (hp.name, cfg[hp.name])
            current = marginal_counts[key]
            target = marginal_targets[key]
            marginal_gain += max(0.0, target - current)
        score += weights["marginal"] * marginal_gain

    if weights["diversity"] > 0:
        if not selected:
            diversity_gain = len(hparams)
        else:
            dmin = min(hamming_distance(cfg, other, names) for other in selected)
            diversity_gain = dmin
        score += weights["diversity"] * diversity_gain

    return score


def summarize_plan(selected, hparams, t_strength, twise_universe, repetitions, used_full_enumeration, full_size, max_unique):
    marginal_counts, twise_counts, twise_presence, min_dist, avg_dist = compute_selected_counts(selected, hparams, t_strength)

    total_possible_twise = sum(len(v) for v in twise_universe.values())
    covered_twise_count = sum(len(twise_presence[k]) for k in twise_presence)
    twise_coverage = 100.0 * covered_twise_count / total_possible_twise if total_possible_twise else 100.0

    marginal_lines = []
    for hp in hparams:
        line = []
        total = sum(marginal_counts[(hp.name, v)] for v in hp.values)
        for v in hp.values:
            c = marginal_counts[(hp.name, v)]
            pct = 100.0 * c / total if total else 0.0
            line.append(f"{v}={c} ({pct:.1f}%)")
        marginal_lines.append(f"{hp.name}: " + ", ".join(line))

    return {
        "unique_experiments": len(selected),
        "total_runs_with_repetitions": len(selected) * repetitions,
        "repetitions": repetitions,
        "full_factorial_size": full_size,
        "used_full_enumeration": used_full_enumeration,
        "enumerated_all_due_to_small_space": used_full_enumeration and full_size <= max_unique,
        "twise_coverage_percent": twise_coverage,
        "t_strength": t_strength,
        "min_hamming_distance": min_dist,
        "avg_hamming_distance": avg_dist,
        "marginals_text": "\n".join(marginal_lines),
    }


def format_seconds(seconds):
    seconds = max(0, float(seconds))
    if seconds < 60:
        return f"{seconds:.1f}s"
    minutes, sec = divmod(seconds, 60)
    if minutes < 60:
        return f"{int(minutes)}m {sec:.1f}s"
    hours, rem = divmod(minutes, 60)
    return f"{int(hours)}h {int(rem)}m {sec:.0f}s"


def plan_experiments(
    hparams,
    max_operations,
    repetitions,
    t_strength,
    twise_weight,
    marginal_weight,
    diversity_weight,
    candidate_pool_size,
    greedy_iterations,
    seed,
    progress_callback=None,
    progress_every=10,
    cancel_check=None,
    refinement_lock_names=None,
    refinement_lock_values=None,
):
    def check_cancel():
        if cancel_check:
            cancel_check()

    if not hparams:
        raise ValueError("Defina pelo menos um hiperparâmetro.")
    for hp in hparams:
        if not hp.name.strip():
            raise ValueError("Todos os hiperparâmetros precisam ter nome.")
        if len(hp.values) < 1:
            raise ValueError(f"O hiperparâmetro '{hp.name}' não possui valores.")
    if repetitions < 1:
        raise ValueError("O número de repetições deve ser >= 1.")
    if max_operations < 1:
        raise ValueError("O máximo de operações deve ser >= 1.")
    if t_strength < 2:
        raise ValueError("A força de cobertura deve ser 2 ou 3.")
    if t_strength > len(hparams):
        t_strength = len(hparams)

    if refinement_lock_names and refinement_lock_values:
        filtered = []
        for hp in hparams:
            if hp.name in refinement_lock_names:
                idx = refinement_lock_names.index(hp.name)
                filtered.append(Hyperparameter(hp.name, [refinement_lock_values[idx]]))
            else:
                filtered.append(hp)
        hparams = filtered
        if progress_callback:
            progress_callback({"type": "log", "message": "Refinement aplicado: alguns hiperparâmetros foram fixados.\n"})

    full_size = full_factorial_count(hparams)
    max_unique = max_operations // repetitions
    if max_unique < 1:
        raise ValueError("Com esse orçamento e esse número de repetições, não cabe nem 1 configuração única.")
    twise_universe = build_twise_universe(hparams, t_strength)

    if progress_callback:
        progress_callback({
            "type": "log",
            "message": (
                f"Início do planejamento.\n"
                f"Hiperparâmetros: {len(hparams)}\n"
                f"Tamanho do espaço completo: {full_size}\n"
                f"Máximo de configurações únicas pelo orçamento: {max_unique}\n"
                f"Força de cobertura: {t_strength}-wise\n"
            )
        })

    check_cancel()

    if full_size <= max_unique:
        if progress_callback:
            progress_callback({"type": "log", "message": "O espaço completo cabe no orçamento. Gerando fatorial completo.\n"})
            progress_callback({"type": "progress", "current": 1, "total": 1})
        selected = generate_full_candidates(hparams)
        summary = summarize_plan(selected, hparams, t_strength, twise_universe, repetitions, True, full_size, max_unique)
        summary["message"] = (
            "O espaço completo cabe no orçamento. "
            "Logo, a cobertura ótima é executar todos os experimentos do fatorial completo."
        )
        summary["first_round_time_seconds"] = 0.0
        summary["avg_round_time_seconds"] = 0.0
        summary["total_heuristic_time_seconds"] = 0.0
        return selected, summary

    rng = random.Random(seed)
    if progress_callback:
        progress_callback({"type": "log", "message": "Gerando pool de candidatos.\n"})
    candidates, used_full_enumeration = generate_candidate_pool(hparams, candidate_pool_size, rng, cancel_check=check_cancel)
    names = [hp.name for hp in hparams]
    candidates = deduplicate_configs(candidates, names)
    if progress_callback:
        progress_callback({"type": "log", "message": f"Pool de candidatos gerado com {len(candidates)} configurações.\n"})
        progress_callback({"type": "progress", "current": 0, "total": max(1, greedy_iterations)})

    weights = {
        "twise": float(twise_weight),
        "marginal": float(marginal_weight),
        "diversity": float(diversity_weight),
    }

    unique_experiments = min(max_unique, len(candidates))
    best_selected = []
    best_tuple = (-1.0, -1.0, -1.0, -1.0)

    total_start = time.perf_counter()
    first_round_time = None
    avg_round_time = None

    for iteration in range(max(1, greedy_iterations)):
        check_cancel()
        round_start = time.perf_counter()
        local_rng = random.Random(seed + iteration)
        shuffled_candidates = list(candidates)
        local_rng.shuffle(shuffled_candidates)

        selected = []
        marginal_counts = Counter()
        twise_presence = defaultdict(set)
        marginal_targets = compute_targets(hparams, unique_experiments)

        available = list(shuffled_candidates)

        while len(selected) < unique_experiments and available:
            check_cancel()
            best_idx = None
            best_score = None
            probe_count = min(len(available), 1000 if len(available) > 1000 else len(available))
            probe = available[:probe_count]
            for idx, cfg in enumerate(probe):
                sc = score_candidate(
                    cfg,
                    selected,
                    hparams,
                    t_strength,
                    twise_presence,
                    marginal_counts,
                    marginal_targets,
                    weights,
                )
                if best_score is None or sc > best_score:
                    best_score = sc
                    best_idx = idx

            chosen = probe[best_idx]
            selected.append(chosen)

            for hp in hparams:
                marginal_counts[(hp.name, chosen[hp.name])] += 1
            for tw_key, tw_val in covered_twise_of_config(chosen, hparams, t_strength):
                twise_presence[tw_key].add(tw_val)

            chosen_key = tuple(chosen[name] for name in names)
            available = [cfg for cfg in available if tuple(cfg[name] for name in names) != chosen_key]
            local_rng.shuffle(available)

        marginal_counts2, twise_counts2, twise_presence2, min_dist, avg_dist = compute_selected_counts(selected, hparams, t_strength)
        total_possible_twise = sum(len(v) for v in twise_universe.values())
        covered_twise_count = sum(len(twise_presence2[k]) for k in twise_presence2)
        twise_cov = covered_twise_count / total_possible_twise if total_possible_twise else 1.0

        imbalance_penalty = 0.0
        for hp in hparams:
            total = sum(marginal_counts2[(hp.name, v)] for v in hp.values)
            target = total / max(1, len(hp.values))
            for v in hp.values:
                imbalance_penalty += abs(marginal_counts2[(hp.name, v)] - target)

        quality_tuple = (twise_cov, -imbalance_penalty, min_dist, avg_dist)
        improved = False
        if quality_tuple > best_tuple:
            best_tuple = quality_tuple
            best_selected = selected
            improved = True

        round_time = time.perf_counter() - round_start
        if iteration == 0:
            first_round_time = round_time
            avg_round_time = round_time
            remaining = max(0, greedy_iterations - 1) * first_round_time
            if progress_callback:
                progress_callback({
                    "type": "log",
                    "message": (
                        f"Primeira rodada concluída em {format_seconds(first_round_time)}. "
                        f"Estimativa inicial para as demais {greedy_iterations - 1} rodadas: {format_seconds(remaining)}.\n"
                    )
                })
        else:
            avg_round_time = ((avg_round_time * iteration) + round_time) / (iteration + 1)

        should_report = ((iteration + 1) % max(1, progress_every) == 0) or (iteration == greedy_iterations - 1)
        if progress_callback:
            progress_callback({"type": "progress", "current": iteration + 1, "total": max(1, greedy_iterations)})
        if should_report and progress_callback:
            elapsed_total = time.perf_counter() - total_start
            remaining_rounds = greedy_iterations - (iteration + 1)
            eta = remaining_rounds * avg_round_time
            progress_callback({
                "type": "log",
                "message": (
                    f"Rodada {iteration + 1}/{greedy_iterations} concluída. "
                    f"Tempo desta rodada: {format_seconds(round_time)}. "
                    f"Tempo médio: {format_seconds(avg_round_time)}. "
                    f"Tempo decorrido: {format_seconds(elapsed_total)}. "
                    f"Estimativa restante: {format_seconds(eta)}. "
                    f"Cobertura parcial {t_strength}-wise do melhor plano: {best_tuple[0] * 100:.2f}%. "
                    f"{'Houve melhora no melhor plano.' if improved else 'Sem melhora no melhor plano.'}\n"
                )
            })

    summary = summarize_plan(best_selected, hparams, t_strength, twise_universe, repetitions, used_full_enumeration, full_size, max_unique)
    summary["message"] = (
        "O espaço completo excede o orçamento. "
        f"Foi gerado um plano aproximado, otimizando cobertura {t_strength}-wise, balanceamento marginal e diversidade."
    )
    summary["max_unique_allowed_by_budget"] = max_unique
    summary["candidate_pool_size"] = len(candidates)
    summary["first_round_time_seconds"] = first_round_time or 0.0
    summary["avg_round_time_seconds"] = avg_round_time or 0.0
    summary["total_heuristic_time_seconds"] = time.perf_counter() - total_start
    return best_selected, summary


def export_to_xlsx(path, experiments, repetitions, summary, phase_name="Plano"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Experimentos"

    if not experiments:
        raise ValueError("Não há experimentos para exportar.")

    names = list(experiments[0].keys())
    headers = names + ["repetitions"]

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for r, cfg in enumerate(experiments, start=2):
        for c, name in enumerate(names, start=1):
            ws.cell(row=r, column=c, value=str(cfg[name]))
        ws.cell(row=r, column=len(names) + 1, value=int(repetitions))

    widths = {name: max(len(name) + 2, 14) for name in headers}
    for cfg in experiments:
        for name in names:
            widths[name] = max(widths[name], len(str(cfg[name])) + 2)
    for i, name in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = min(max(widths[name], 14), 40)

    meta = wb.create_sheet("Resumo")
    meta["A1"] = "Resumo do Planejamento"
    meta["A1"].font = Font(bold=True, size=13)

    items = [
        ("Fase", phase_name),
        ("Mensagem", summary.get("message", "")),
        ("Tamanho do fatorial completo", summary.get("full_factorial_size", "")),
        ("Configurações únicas selecionadas", summary.get("unique_experiments", "")),
        ("Repetições por configuração", summary.get("repetitions", "")),
        ("Execuções totais", summary.get("total_runs_with_repetitions", "")),
        (f"Cobertura {summary.get('t_strength', 2)}-wise (%)", round(summary.get("twise_coverage_percent", 0.0), 3)),
        ("Distância mínima de Hamming", summary.get("min_hamming_distance", "")),
        ("Distância média de Hamming", round(summary.get("avg_hamming_distance", 0.0), 3)),
        ("Tempo da primeira rodada", format_seconds(summary.get("first_round_time_seconds", 0.0))),
        ("Tempo médio por rodada", format_seconds(summary.get("avg_round_time_seconds", 0.0))),
        ("Tempo total da heurística", format_seconds(summary.get("total_heuristic_time_seconds", 0.0))),
        ("Distribuições marginais", summary.get("marginals_text", "")),
    ]
    row = 3
    for key, value in items:
        meta.cell(row=row, column=1, value=key).font = Font(bold=True)
        meta.cell(row=row, column=2, value=str(value))
        row += 1

    meta.column_dimensions["A"].width = 36
    meta.column_dimensions["B"].width = 100
    wb.save(path)


HELP_TEXT = """
Esta versão executa a heurística em thread separada, com:

1. Log incremental na tela.
2. Barra de progresso.
3. Botão de cancelamento.
4. Medição do tempo da primeira rodada.
5. Estimativa de tempo restante.
6. Atualização do log a cada N rodadas, configurável.

Como usar
---------
1. Defina os hiperparâmetros.
2. Ajuste o orçamento, repetições e cobertura.
3. Escolha o intervalo de atualização do log.
4. Execute screening ou refinement.
5. Se necessário, cancele a execução.

Critérios
---------
- Cobertura marginal: balanceia valores por hiperparâmetro.
- Cobertura 2-wise: cobre pares de hiperparâmetros.
- Cobertura 3-wise: cobre trincas.
- Diversidade: espalha configurações usando distância de Hamming.

Observação
----------
Se o fatorial completo cabe no orçamento, a solução é exata.
Caso contrário, a solução é heurística aproximada.
"""


class HyperparameterTable(ttk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self._build()

    def _build(self):
        tools = ttk.Frame(self)
        tools.pack(fill="x", pady=(0, 6))
        ttk.Button(tools, text="Adicionar hiperparâmetro", command=self.add_row).pack(side="left")
        ttk.Button(tools, text="Remover selecionado", command=self.remove_selected).pack(side="left", padx=6)
        ttk.Button(tools, text="Limpar tudo", command=self.clear_rows).pack(side="left", padx=6)

        cols = ("name", "values")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", selectmode="browse", height=12)
        self.tree.heading("name", text="Hiperparâmetro")
        self.tree.heading("values", text="Valores possíveis, separados por vírgula ou ;")
        self.tree.column("name", width=240, anchor="w")
        self.tree.column("values", width=720, anchor="w")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        form = ttk.LabelFrame(self, text="Editor", padding=8)
        form.pack(fill="x", pady=(8, 0))
        ttk.Label(form, text="Nome").grid(row=0, column=0, sticky="w")
        ttk.Label(form, text="Valores").grid(row=1, column=0, sticky="w")
        self.name_var = tk.StringVar()
        self.values_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.name_var, width=30).grid(row=0, column=1, sticky="ew", padx=6, pady=3)
        ttk.Entry(form, textvariable=self.values_var, width=90).grid(row=1, column=1, sticky="ew", padx=6, pady=3)
        btns = ttk.Frame(form)
        btns.grid(row=2, column=1, sticky="w", pady=(6, 0))
        ttk.Button(btns, text="Salvar linha", command=self.save_current).pack(side="left")
        ttk.Button(btns, text="Nova linha", command=lambda: self._prepare_new()).pack(side="left", padx=6)
        form.columnconfigure(1, weight=1)
        self.current_item = None
        self._prepare_new()

    def _prepare_new(self):
        self.current_item = None
        self.name_var.set("")
        self.values_var.set("")

    def on_select(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        item = selected[0]
        values = self.tree.item(item, "values")
        self.current_item = item
        self.name_var.set(values[0])
        self.values_var.set(values[1])

    def save_current(self):
        name = self.name_var.get().strip()
        values_text = self.values_var.get().strip()
        if not name:
            messagebox.showerror("Erro", "Informe o nome do hiperparâmetro.")
            return
        values = normalize_values_text(values_text)
        if not values:
            messagebox.showerror("Erro", "Informe ao menos um valor possível.")
            return
        pretty_values = ", ".join(values)
        if self.current_item is None:
            self.tree.insert("", "end", values=(name, pretty_values))
        else:
            self.tree.item(self.current_item, values=(name, pretty_values))
        self._prepare_new()

    def add_row(self, name="", values=""):
        self.tree.insert("", "end", values=(name, values))

    def remove_selected(self):
        selected = self.tree.selection()
        if selected:
            self.tree.delete(selected[0])
            self._prepare_new()

    def clear_rows(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._prepare_new()

    def get_hyperparameters(self):
        hps = []
        for item in self.tree.get_children():
            name, values_text = self.tree.item(item, "values")
            name = name.strip()
            values = normalize_values_text(values_text)
            if not name:
                raise ValueError("Há um hiperparâmetro sem nome.")
            if not values:
                raise ValueError(f"O hiperparâmetro '{name}' não possui valores.")
            hps.append(Hyperparameter(name=name, values=values))
        return hps

    def set_hyperparameters(self, hparams):
        self.clear_rows()
        for hp in hparams:
            self.tree.insert("", "end", values=(hp["name"], ", ".join(hp["values"])))


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1280x930")
        self.minsize(1080, 780)

        self.last_result = None
        self.last_refinement_result = None
        self.worker_thread = None
        self.worker_queue = queue.Queue()
        self.cancel_event = threading.Event()
        self.running = False

        self._build_ui()
        self.after(100, self.process_worker_queue)

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        top = ttk.Frame(self, padding=10)
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(0, weight=1)

        title = ttk.Label(top, text=APP_TITLE, font=("TkDefaultFont", 14, "bold"))
        title.grid(row=0, column=0, sticky="w")

        toolbar = ttk.Frame(top)
        toolbar.grid(row=0, column=1, sticky="e")
        self.btn_new = ttk.Button(toolbar, text="Novo", command=self.new_config)
        self.btn_open = ttk.Button(toolbar, text="Abrir JSON", command=self.load_json)
        self.btn_save = ttk.Button(toolbar, text="Salvar JSON", command=self.save_json)
        self.btn_screen = ttk.Button(toolbar, text="Executar screening", command=self.run_screening)
        self.btn_refine = ttk.Button(toolbar, text="Executar refinement", command=self.run_refinement)
        self.btn_cancel = ttk.Button(toolbar, text="Cancelar", command=self.cancel_run, state="disabled")
        self.btn_export = ttk.Button(toolbar, text="Exportar XLSX", command=self.export_xlsx)
        self.btn_help = ttk.Button(toolbar, text="Help", command=self.show_help)
        for btn in [self.btn_new, self.btn_open, self.btn_save, self.btn_screen, self.btn_refine, self.btn_cancel, self.btn_export, self.btn_help]:
            btn.pack(side="left", padx=3)

        notebook = ttk.Notebook(self)
        notebook.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        self.notebook = notebook

        tab_hp = ttk.Frame(notebook, padding=10)
        tab_plan = ttk.Frame(notebook, padding=10)
        tab_phase = ttk.Frame(notebook, padding=10)
        tab_result = ttk.Frame(notebook, padding=10)

        notebook.add(tab_hp, text="Hiperparâmetros")
        notebook.add(tab_plan, text="Planejamento")
        notebook.add(tab_phase, text="Fases")
        notebook.add(tab_result, text="Resultado")

        tab_hp.columnconfigure(0, weight=1)
        tab_hp.rowconfigure(0, weight=1)
        self.hp_table = HyperparameterTable(tab_hp)
        self.hp_table.grid(row=0, column=0, sticky="nsew")

        self.max_operations_var = tk.IntVar(value=100)
        self.repetitions_var = tk.IntVar(value=3)
        self.t_strength_var = tk.IntVar(value=2)
        self.twise_weight_var = tk.DoubleVar(value=10.0)
        self.marginal_weight_var = tk.DoubleVar(value=3.0)
        self.diversity_weight_var = tk.DoubleVar(value=1.0)
        self.candidate_pool_var = tk.IntVar(value=DEFAULT_CANDIDATE_POOL)
        self.greedy_iterations_var = tk.IntVar(value=DEFAULT_ITERATIONS)
        self.progress_every_var = tk.IntVar(value=10)
        self.seed_var = tk.IntVar(value=RANDOM_SEED)

        frm = ttk.LabelFrame(tab_plan, text="Configuração do planejamento", padding=10)
        frm.pack(fill="x", anchor="n")
        plan_fields = [
            ("Máximo de operações", self.max_operations_var),
            ("Repetições por configuração", self.repetitions_var),
            ("Força de cobertura t (2 ou 3)", self.t_strength_var),
            ("Peso da cobertura t-wise", self.twise_weight_var),
            ("Peso da cobertura marginal", self.marginal_weight_var),
            ("Peso da diversidade", self.diversity_weight_var),
            ("Tamanho do pool de candidatos", self.candidate_pool_var),
            ("Iterações da heurística", self.greedy_iterations_var),
            ("Atualizar log a cada N rodadas", self.progress_every_var),
            ("Semente aleatória", self.seed_var),
        ]
        for row, (label, var) in enumerate(plan_fields):
            ttk.Label(frm, text=label).grid(row=row, column=0, sticky="w", pady=4)
            ttk.Entry(frm, textvariable=var, width=18).grid(row=row, column=1, sticky="w", pady=4, padx=8)

        ttk.Label(
            frm,
            text=(
                "A primeira rodada é cronometrada separadamente.\n"
                "Depois o programa estima o tempo restante e atualiza o log periodicamente."
            )
        ).grid(row=len(plan_fields), column=0, columnspan=2, sticky="w", pady=(10, 0))

        phase_outer = ttk.Frame(tab_phase)
        phase_outer.pack(fill="both", expand=True)
        phase_outer.columnconfigure(0, weight=1)
        phase_outer.columnconfigure(1, weight=1)
        phase_outer.rowconfigure(1, weight=1)

        descr = ttk.Label(
            phase_outer,
            text=(
                "Screening explora o espaço amplo. Refinement permite fixar alguns hiperparâmetros\n"
                "em valores escolhidos e gerar um novo plano para a sub-região considerada mais promissora."
            )
        )
        descr.grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        lock_frame = ttk.LabelFrame(phase_outer, text="Fixação de hiperparâmetros para refinement", padding=8)
        lock_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 8))
        lock_frame.columnconfigure(0, weight=1)
        lock_frame.rowconfigure(0, weight=1)

        self.lock_tree = ttk.Treeview(lock_frame, columns=("name", "value"), show="headings", height=18)
        self.lock_tree.heading("name", text="Hiperparâmetro")
        self.lock_tree.heading("value", text="Valor fixado")
        self.lock_tree.column("name", width=220, anchor="w")
        self.lock_tree.column("value", width=220, anchor="w")
        self.lock_tree.grid(row=0, column=0, sticky="nsew")

        lock_buttons = ttk.Frame(lock_frame)
        lock_buttons.grid(row=1, column=0, sticky="w", pady=8)
        ttk.Button(lock_buttons, text="Atualizar lista", command=self.refresh_lock_table).pack(side="left")
        ttk.Button(lock_buttons, text="Editar seleção", command=self.edit_lock_selection).pack(side="left", padx=6)
        ttk.Button(lock_buttons, text="Limpar fixações", command=self.clear_locks).pack(side="left")

        side = ttk.LabelFrame(phase_outer, text="Ações", padding=8)
        side.grid(row=1, column=1, sticky="nsew")
        ttk.Button(side, text="Executar screening", command=self.run_screening).pack(fill="x", pady=4)
        ttk.Button(side, text="Copiar primeiro experimento do screening para fixações", command=self.seed_refinement_from_best).pack(fill="x", pady=4)
        ttk.Button(side, text="Executar refinement", command=self.run_refinement).pack(fill="x", pady=4)

        result_frame = ttk.Frame(tab_result)
        result_frame.pack(fill="both", expand=True)
        result_frame.rowconfigure(1, weight=1)
        result_frame.columnconfigure(0, weight=1)

        status_frame = ttk.Frame(result_frame)
        status_frame.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        status_frame.columnconfigure(1, weight=1)

        ttk.Label(status_frame, text="Status").grid(row=0, column=0, sticky="w")
        self.status_var = tk.StringVar(value="Pronto")
        ttk.Label(status_frame, textvariable=self.status_var).grid(row=0, column=1, sticky="w", padx=8)

        self.progress = ttk.Progressbar(status_frame, orient="horizontal", mode="determinate")
        self.progress.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(6, 0))

        self.output = tk.Text(result_frame, wrap="word", font=("Consolas", 10))
        self.output.grid(row=1, column=0, sticky="nsew")

        self.bootstrap_defaults()
        self.append_output("Pronto.\nDefina os hiperparâmetros e execute o screening.\n")

    def bootstrap_defaults(self):
        if not self.hp_table.tree.get_children():
            example = [
                {"name": "modelo", "values": ["A", "B", "C"]},
                {"name": "batch", "values": ["16", "32"]},
                {"name": "optimizer", "values": ["adam", "sgd"]},
            ]
            self.hp_table.set_hyperparameters(example)
        self.refresh_lock_table()

    def append_output(self, text):
        self.output.insert("end", text)
        self.output.see("end")

    def clear_output(self):
        self.output.delete("1.0", "end")

    def set_running_state(self, running):
        self.running = running
        state_main = "disabled" if running else "normal"
        self.btn_new.configure(state=state_main)
        self.btn_open.configure(state=state_main)
        self.btn_save.configure(state=state_main)
        self.btn_screen.configure(state=state_main)
        self.btn_refine.configure(state=state_main)
        self.btn_cancel.configure(state="normal" if running else "disabled")
        self.btn_export.configure(state="disabled" if running else "normal")
        self.btn_help.configure(state=state_main)

    def refresh_lock_table(self):
        existing = {}
        for item in self.lock_tree.get_children():
            name, value = self.lock_tree.item(item, "values")
            existing[name] = value
        for item in self.lock_tree.get_children():
            self.lock_tree.delete(item)
        for hp in self.hp_table.get_hyperparameters():
            value = existing.get(hp.name, "")
            self.lock_tree.insert("", "end", values=(hp.name, value))

    def edit_lock_selection(self):
        sel = self.lock_tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Selecione uma linha de fixação.")
            return
        item = sel[0]
        name, current_value = self.lock_tree.item(item, "values")
        hps = self.hp_table.get_hyperparameters()
        values = None
        for hp in hps:
            if hp.name == name:
                values = hp.values
                break
        if values is None:
            return

        win = tk.Toplevel(self)
        win.title(f"Fixar valor para {name}")
        win.geometry("360x160")
        ttk.Label(win, text=f"Hiperparâmetro: {name}").pack(anchor="w", padx=10, pady=(12, 6))
        var = tk.StringVar(value=current_value if current_value in values else "")
        combo = ttk.Combobox(win, textvariable=var, values=[""] + values, state="readonly")
        combo.pack(fill="x", padx=10)
        ttk.Label(win, text="Valor vazio significa não fixar este hiperparâmetro.").pack(anchor="w", padx=10, pady=6)

        def save():
            self.lock_tree.item(item, values=(name, var.get()))
            win.destroy()

        ttk.Button(win, text="Salvar", command=save).pack(pady=10)

    def clear_locks(self):
        for item in self.lock_tree.get_children():
            name, _ = self.lock_tree.item(item, "values")
            self.lock_tree.item(item, values=(name, ""))

    def seed_refinement_from_best(self):
        if not self.last_result or not self.last_result["experiments"]:
            messagebox.showinfo("Info", "Execute o screening antes.")
            return
        best = self.last_result["experiments"][0]
        for item in self.lock_tree.get_children():
            name, _ = self.lock_tree.item(item, "values")
            self.lock_tree.item(item, values=(name, best.get(name, "")))
        messagebox.showinfo("Info", "As fixações foram preenchidas com o primeiro experimento do screening.\nRevise antes de rodar refinement.")

    def get_locks(self):
        names = []
        values = []
        for item in self.lock_tree.get_children():
            name, value = self.lock_tree.item(item, "values")
            if value:
                names.append(name)
                values.append(value)
        return names, values

    def get_config_dict(self):
        hparams = [asdict(hp) for hp in self.hp_table.get_hyperparameters()]
        lock_names, lock_values = self.get_locks()
        return {
            "hyperparameters": hparams,
            "max_operations": self.max_operations_var.get(),
            "repetitions": self.repetitions_var.get(),
            "t_strength": self.t_strength_var.get(),
            "weights": {
                "twise": self.twise_weight_var.get(),
                "marginal": self.marginal_weight_var.get(),
                "diversity": self.diversity_weight_var.get(),
            },
            "candidate_pool_size": self.candidate_pool_var.get(),
            "greedy_iterations": self.greedy_iterations_var.get(),
            "progress_every": self.progress_every_var.get(),
            "seed": self.seed_var.get(),
            "refinement_locks": {
                "names": lock_names,
                "values": lock_values,
            },
        }

    def new_config(self):
        self.hp_table.clear_rows()
        self.max_operations_var.set(100)
        self.repetitions_var.set(3)
        self.t_strength_var.set(2)
        self.twise_weight_var.set(10.0)
        self.marginal_weight_var.set(3.0)
        self.diversity_weight_var.set(1.0)
        self.candidate_pool_var.set(DEFAULT_CANDIDATE_POOL)
        self.greedy_iterations_var.set(DEFAULT_ITERATIONS)
        self.progress_every_var.set(10)
        self.seed_var.set(RANDOM_SEED)
        self.last_result = None
        self.last_refinement_result = None
        self.bootstrap_defaults()
        self.clear_output()
        self.status_var.set("Pronto")
        self.progress.configure(value=0, maximum=1)
        self.append_output("Nova configuração criada.\n")

    def save_json(self):
        try:
            data = self.get_config_dict()
        except Exception as e:
            messagebox.showerror("Erro", str(e))
            return
        path = filedialog.asksaveasfilename(
            title="Salvar configuração",
            defaultextension=".json",
            filetypes=[("JSON", "*.json")]
        )
        if not path:
            return
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self.append_output(f"Configuração salva em: {path}\n")

    def load_json(self):
        path = filedialog.askopenfilename(
            title="Abrir configuração",
            filetypes=[("JSON", "*.json")]
        )
        if not path:
            return
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        self.hp_table.set_hyperparameters(data.get("hyperparameters", []))
        self.max_operations_var.set(int(data.get("max_operations", 100)))
        self.repetitions_var.set(int(data.get("repetitions", 3)))
        self.t_strength_var.set(int(data.get("t_strength", 2)))
        weights = data.get("weights", {})
        self.twise_weight_var.set(float(weights.get("twise", 10.0)))
        self.marginal_weight_var.set(float(weights.get("marginal", 3.0)))
        self.diversity_weight_var.set(float(weights.get("diversity", 1.0)))
        self.candidate_pool_var.set(int(data.get("candidate_pool_size", DEFAULT_CANDIDATE_POOL)))
        self.greedy_iterations_var.set(int(data.get("greedy_iterations", DEFAULT_ITERATIONS)))
        self.progress_every_var.set(int(data.get("progress_every", 10)))
        self.seed_var.set(int(data.get("seed", RANDOM_SEED)))
        self.refresh_lock_table()
        locks = data.get("refinement_locks", {})
        names = locks.get("names", [])
        values = locks.get("values", [])
        lock_map = dict(zip(names, values))
        for item in self.lock_tree.get_children():
            name, _ = self.lock_tree.item(item, "values")
            self.lock_tree.item(item, values=(name, lock_map.get(name, "")))
        self.last_result = None
        self.last_refinement_result = None
        self.clear_output()
        self.append_output(f"Configuração carregada de: {path}\n")

    def show_help(self):
        win = tk.Toplevel(self)
        win.title("Help")
        win.geometry("980x720")
        text = tk.Text(win, wrap="word")
        text.pack(fill="both", expand=True)
        text.insert("1.0", HELP_TEXT)
        text.configure(state="disabled")

    def _format_result(self, experiments, summary, phase_name):
        lines = []
        lines.append(f"Fase: {phase_name}")
        lines.append(summary.get("message", ""))
        lines.append("")
        lines.append(f"Tamanho do espaço completo: {summary['full_factorial_size']}")
        lines.append(f"Configurações únicas selecionadas: {summary['unique_experiments']}")
        lines.append(f"Repetições por configuração: {summary['repetitions']}")
        lines.append(f"Total de execuções: {summary['total_runs_with_repetitions']}")
        lines.append(f"Cobertura {summary.get('t_strength', 2)}-wise estimada: {summary['twise_coverage_percent']:.2f}%")
        lines.append(f"Distância mínima de Hamming: {summary['min_hamming_distance']}")
        lines.append(f"Distância média de Hamming: {summary['avg_hamming_distance']:.2f}")
        lines.append(f"Tempo da primeira rodada: {format_seconds(summary.get('first_round_time_seconds', 0.0))}")
        lines.append(f"Tempo médio por rodada: {format_seconds(summary.get('avg_round_time_seconds', 0.0))}")
        lines.append(f"Tempo total da heurística: {format_seconds(summary.get('total_heuristic_time_seconds', 0.0))}")
        if "max_unique_allowed_by_budget" in summary:
            lines.append(f"Máximo de configurações únicas permitido pelo orçamento: {summary['max_unique_allowed_by_budget']}")
        if "candidate_pool_size" in summary:
            lines.append(f"Tamanho efetivo do pool de candidatos: {summary['candidate_pool_size']}")
        lines.append("")
        lines.append("Distribuições marginais:")
        lines.append(summary["marginals_text"])
        lines.append("")
        lines.append("Experimentos propostos:")
        names = list(experiments[0].keys()) if experiments else []
        if names:
            header = " | ".join(names) + " | repetitions"
            lines.append(header)
            lines.append("-" * min(220, len(header) + 10))
        for cfg in experiments:
            row = " | ".join(str(cfg[name]) for name in names) + f" | {summary['repetitions']}"
            lines.append(row)
        lines.append("\n")
        return "\n".join(lines)

    def start_worker(self, phase_name, refinement=False):
        if self.running:
            return
        try:
            cfg = self.get_config_dict()
            hparams = [Hyperparameter(**hp) for hp in cfg["hyperparameters"]]
        except Exception as e:
            messagebox.showerror("Erro", str(e))
            return

        lock_names, lock_values = [], []
        if refinement:
            locks = cfg.get("refinement_locks", {})
            lock_names = locks.get("names", [])
            lock_values = locks.get("values", [])
            if not lock_names:
                if not messagebox.askyesno("Refinement", "Nenhum hiperparâmetro foi fixado.\nDeseja executar refinement assim mesmo?"):
                    return

        self.clear_output()
        self.notebook.select(3)
        self.append_output(f"Executando {phase_name.lower()}.\n")
        self.status_var.set(f"Executando {phase_name.lower()}...")
        self.progress.configure(value=0, maximum=max(1, int(cfg["greedy_iterations"])))
        self.cancel_event.clear()
        self.set_running_state(True)

        def progress_callback(payload):
            self.worker_queue.put(("progress", payload))

        def cancel_check():
            if self.cancel_event.is_set():
                raise CancelledError("Execução cancelada pelo usuário.")

        def worker():
            try:
                experiments, summary = plan_experiments(
                    hparams=hparams,
                    max_operations=int(cfg["max_operations"]),
                    repetitions=int(cfg["repetitions"]),
                    t_strength=int(cfg["t_strength"]),
                    twise_weight=float(cfg["weights"]["twise"]),
                    marginal_weight=float(cfg["weights"]["marginal"]),
                    diversity_weight=float(cfg["weights"]["diversity"]),
                    candidate_pool_size=int(cfg["candidate_pool_size"]),
                    greedy_iterations=int(cfg["greedy_iterations"]),
                    seed=int(cfg["seed"]) + (1000 if refinement else 0),
                    progress_callback=progress_callback,
                    progress_every=int(cfg["progress_every"]),
                    cancel_check=cancel_check,
                    refinement_lock_names=lock_names,
                    refinement_lock_values=lock_values,
                )
                self.worker_queue.put(("done", {
                    "experiments": experiments,
                    "summary": summary,
                    "config": cfg,
                    "phase": phase_name,
                    "refinement": refinement,
                }))
            except CancelledError as e:
                self.worker_queue.put(("cancelled", str(e)))
            except Exception as e:
                self.worker_queue.put(("error", str(e)))

        self.worker_thread = threading.Thread(target=worker, daemon=True)
        self.worker_thread.start()

    def run_screening(self):
        self.start_worker("Screening", refinement=False)

    def run_refinement(self):
        self.start_worker("Refinement", refinement=True)

    def cancel_run(self):
        if self.running:
            self.cancel_event.set()
            self.status_var.set("Cancelando...")

    def process_worker_queue(self):
        try:
            while True:
                msg_type, payload = self.worker_queue.get_nowait()
                if msg_type == "progress":
                    if payload["type"] == "log":
                        self.append_output(payload["message"])
                    elif payload["type"] == "progress":
                        self.progress.configure(maximum=max(1, payload["total"]))
                        self.progress.configure(value=payload["current"])
                        self.status_var.set(f"Executando... rodada {payload['current']}/{payload['total']}")
                elif msg_type == "done":
                    if payload["refinement"]:
                        self.last_refinement_result = payload
                    else:
                        self.last_result = payload
                    self.append_output("\nResumo final\n")
                    self.append_output(self._format_result(payload["experiments"], payload["summary"], payload["phase"]))
                    self.status_var.set(f"{payload['phase']} concluído")
                    self.set_running_state(False)
                elif msg_type == "cancelled":
                    self.append_output(f"\nExecução cancelada.\n{payload}\n")
                    self.status_var.set("Execução cancelada")
                    self.set_running_state(False)
                elif msg_type == "error":
                    self.append_output(f"\nErro: {payload}\n")
                    self.status_var.set("Erro")
                    self.set_running_state(False)
                    messagebox.showerror("Erro", payload)
        except queue.Empty:
            pass
        self.after(100, self.process_worker_queue)

    def export_xlsx(self):
        result = self.last_refinement_result or self.last_result
        if not result:
            messagebox.showwarning("Aviso", "Execute screening ou refinement antes de exportar.")
            return
        path = filedialog.asksaveasfilename(
            title="Exportar planilha",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return
        export_to_xlsx(
            path=path,
            experiments=result["experiments"],
            repetitions=result["summary"]["repetitions"],
            summary=result["summary"],
            phase_name=result["phase"],
        )
        self.append_output(f"\nPlanilha exportada para: {path}\n")


if __name__ == "__main__":
    app = App()
    app.mainloop()
